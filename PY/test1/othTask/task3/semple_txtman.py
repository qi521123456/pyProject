# coding=utf-8

from openpyxl import load_workbook,workbook,worksheet
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

for num in range(1,53):
    i=2
    iflag=False
    pflag=False
    filename = 'D:/fofa_result/Result_fingerprint['+str(num)+'].txt'
    print(filename)
    wb=Workbook()
    ws=wb.active
    #sheet = wb.get_sheet_by_name('Sheet1')
    #ew=ExcelWriter(workbook=wb)
    #ws = wb.worksheets[0]
    #sheet.title='te'
    #ew.save(wb)


    with open(filename,'r',encoding='utf-8') as fileRead:
        lines = fileRead.readlines()  # 读取全部内容
        for line in lines:
            if not line:
                break
            if iflag:
                #print(line+'------'+'------'+str(i))
                iflag=False
                ws['A'+str(i)] = line.strip()
                #ws.append([1, 2, 3])
            if pflag:
                ws['B' + str(i)] = line.strip()
                pflag = False
            if line.strip() == 'Auto_port:':
                pflag=True
            if line.strip() == 'Auto_ip:':
                iflag=True
            if line.strip() == 'HTTP_Head:':
                i+=1
                #print('------' + str(i))



    #wb.save('D:/xlstest/test_'+str(num)+'.xlsx')
    fileRead.close()
    #print(num)