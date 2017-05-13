'''
Created on 2016年5月25日

@author: lihongbiao
'''
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill, FILL_SOLID
    
def create_sheet(filename):
    sheetname = 'testsheet'
    wb = load_workbook(filename)
    wb.create_sheet(sheetname)
    ws = wb.get_sheet_by_name(sheetname)
    ws.merge_cells(start_row=1,start_column=2,end_row=1,end_column=7)
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25 
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25

    ft = Font(name='Courier New', size=16, color=colors.WHITE, bold=True, ) #字体
    fill = PatternFill(fill_type=FILL_SOLID, start_color='003366FF') #背景色
    alig = Alignment(horizontal='center', vertical='center') #对齐方式

    ws['B1'].fill = fill
    ws['B1'].alignment = alig
    ws['B1'].font = ft
    ws['B1'].value = '目标设备'

    ft2 = Font(name='Courier New', size=12, color=colors.BLACK, bold=True,)

    ws['B2'].font = ft2
    ws['C2'].font = ft2
    ws['D2'].font = ft2
    ws['E2'].font = ft2
    ws['F2'].font = ft2
    ws['G2'].font = ft2

    ws['B2'].fill = fill
    ws['C2'].fill = fill
    ws['D2'].fill = fill
    ws['E2'].fill = fill
    ws['F2'].fill = fill
    ws['G2'].fill = fill

    ws['B2'].value = 'IP'
    ws['C2'].value = '位置'
    ws['D2'].value = '设备型号'
    ws['E2'].value = '固件版本号'
    ws['F2'].value = '其他'
    ws['G2'].value = 'VPS'
    wb.save(filename)

def create_file(filename):
    wb = Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1,start_column=2,end_row=1,end_column=7)
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25 
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25

    ft = Font(name='Courier New', size=16, color=colors.WHITE, bold=True, ) #字体
    fill = PatternFill(fill_type=FILL_SOLID, start_color='003366FF') #背景色
    alig = Alignment(horizontal='center', vertical='center') #对齐方式

    ws['B1'].fill = fill
    ws['B1'].alignment = alig
    ws['B1'].font = ft
    ws['B1'].value = '目标设备'

    ft2 = Font(name='Courier New', size=12, color=colors.BLACK, bold=True,)

    ws['B2'].font = ft2
    ws['C2'].font = ft2
    ws['D2'].font = ft2
    ws['E2'].font = ft2
    ws['F2'].font = ft2
    ws['G2'].font = ft2

    ws['B2'].fill = fill
    ws['C2'].fill = fill
    ws['D2'].fill = fill
    ws['E2'].fill = fill
    ws['F2'].fill = fill
    ws['G2'].fill = fill

    ws['B2'].value = 'IP'
    ws['C2'].value = '位置'
    ws['D2'].value = '设备型号'
    ws['E2'].value = '固件版本号'
    ws['F2'].value = '其他'
    ws['G2'].value = 'VPS'
    wb.save(filename)

def main():
    filename = 'D:/styles.xlsx'
    if os.path.exists(filename):
        create_sheet(filename)
    else:
        create_file(filename)

if __name__ == '__main__':
    main()