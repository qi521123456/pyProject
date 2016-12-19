# coding=utf-8
import re
from openpyxl import Workbook

def titleChk(t_str):
    p=re.compile("\[(.+)\]")
    if p.match(t_str):
        return True
    else:
        return False


wb = Workbook()
ws = wb.active

def readFile_writeXlsx(filenum):
    pflag = False
    iflag = False
    count = 2
    filename = 'D:/fofa_result/Result_fingerprint[' + str(filenum) + '].txt'
    with open(filename,'r') as fileRead:
        lines = fileRead.readlines()  # 读取全部内容
        for line in lines:
            if not line:
                break
            if iflag:
                #print(line+'------'+'------'+str(i))
                iflag=False
                ws['A'+str(count)] = line.strip()
            if pflag:
                ws['B' + str(count)] = line.strip()
                pflag = False
            if titleChk(line):
                count += 1
            if line.strip() == 'Auto_port:':
                pflag = True
            if line.strip() == 'Auto_ip:':
                iflag = True
