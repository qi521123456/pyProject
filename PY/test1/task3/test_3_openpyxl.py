# coding=utf-8
from openpyxl import Workbook,load_workbook

def test(filename):
    wb = Workbook()
    # 创建一个sheet
    #ws = wb.get_sheet_by_name('Sheet1')
    # 写入10行,每行20个
    ws = wb.active
    print(wb.get_sheet_names())
    ws['a2'] = 'woe'
    wb.save(filename)
def test2():
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    dest_filename = 'D:/empty_book.xlsx'
    #ws1 = wb.active
    ws1 = wb.create_sheet()
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    wb.save(filename=dest_filename)


if __name__ == '__main__':
    test('D:/myExcelFile.xlsx')
    # test2()



